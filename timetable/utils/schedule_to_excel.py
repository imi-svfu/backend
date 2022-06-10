import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, numbers, Border
from openpyxl.utils import get_column_letter

from timetable import models

app_dir = os.path.dirname(os.path.abspath(__file__))
template = os.path.join(app_dir, 'schedule.xlsx')
workbook = load_workbook(template)
source = workbook.active

styles = {
    'wrap_text': Alignment(horizontal="center",
                           vertical="center",
                           text_rotation=0,
                           wrap_text=True,
                           shrink_to_fit=False,
                           indent=0),
    'default_font': Font(name="Times New Roman",
                         size=14),
    'header_font': Font(name="Times New Roman",
                        size=14,
                        bold="b"),
}

star = {
    0: "",
    1: "*",
    2: "**",
}

schedule_types = {
    "LEC": "Лек",
    "PRA": "Пр",
    "LAB": "Лаб",
}


def fill_borders(ws, groups):
    for i in range(groups.count() - 1):
        for row in ws[get_column_letter(3 + 4 * i) + '3':get_column_letter(6 + 4 * i) + '41']:
            for cell in row:
                dist_cell = ws.cell(column=cell.column + 4, row=cell.row)
                dist_cell.border = Border(**(dist_cell.border.__dict__ | cell.border.__dict__))


def fill_cell(ws, column, j, content):
    return ((str(ws.cell(column=column, row=j).value) + "\n")
            if ws.cell(column=column, row=j).value is not None else "") + content


def export():
    for course in range(4):
        # groups = models.Group.objects.filter(name__contains="-18")
        groups = models.Group.objects.filter(name__contains=f'-{22 - course - 1}')

        if groups.count() > 0:
            ws = workbook.copy_worksheet(source)
            ws.title = f'{course + 1} курс'

            c = 0
            for group in groups:
                c += 1
                semester = models.Semester.objects.get(group__id=group.id)

                current_col_start = int(3 + (c - 1) * 4)

                ws.column_dimensions[get_column_letter(current_col_start)].width = 35
                ws.column_dimensions[get_column_letter(current_col_start + 1)].width = 19
                ws.column_dimensions[get_column_letter(current_col_start + 2)].width = 14
                ws.column_dimensions[get_column_letter(current_col_start + 3)].width = 14

                gdescell = ws.cell(column=current_col_start, row=3)  # group description cell
                gdescell.value = "Наименование группы"
                gdescell.font = styles['default_font']

                scdescell = ws.cell(column=current_col_start + 1, row=3)  # students count description cell
                scdescell.value = "Общее число обучающихся, чел."
                scdescell.font = styles['default_font']

                gcell = ws.cell(column=current_col_start, row=4)  # group name cell
                gcell.value = group.name
                gcell.font = styles['header_font']
                gcell.alignment = styles['wrap_text']

                sbcell = ws.cell(column=current_col_start + 2, row=4)  # semester begin cell
                sbcell.value = semester.study_start
                sbcell.font = styles['header_font']
                sbcell.alignment = styles['wrap_text']
                sbcell.number_format = numbers.BUILTIN_FORMATS[16]

                secell = ws.cell(column=current_col_start + 3, row=4)
                secell.value = semester.study_end
                secell.font = styles['header_font']
                secell.alignment = styles['wrap_text']
                secell.number_format = numbers.BUILTIN_FORMATS[16]

                disdescell = ws.cell(column=current_col_start, row=5)  # discipline desc cell
                disdescell.value = "Дисциплина"
                disdescell.alignment = styles['wrap_text']
                disdescell.font = styles['default_font']

                lecdescell = ws.cell(column=current_col_start + 1, row=5)  # lecturer desc cell
                lecdescell.value = "ФИО преподавателя"
                lecdescell.alignment = styles['wrap_text']
                lecdescell.font = styles['default_font']

                typedescell = ws.cell(column=current_col_start + 2, row=5)  # type desc cell
                typedescell.value = "Вид учебного занятия"
                typedescell.alignment = styles['wrap_text']
                typedescell.font = styles['default_font']

                roomdescell = ws.cell(column=current_col_start + 3, row=5)  # room desc cell
                roomdescell.value = "ауд."
                roomdescell.alignment = styles['wrap_text']
                roomdescell.font = styles['default_font']

                schedules = models.Schedule.objects.filter(lesson__group__id=group.id).order_by("week_day", "pair_num")

                fill_borders(ws, groups)

                for i in range(5, 42):
                    dis_cell = ws.cell(column=current_col_start, row=i)  # discipline cell
                    dis_cell.alignment = styles['wrap_text']
                    dis_cell.font = styles['default_font']

                    lec_cell = ws.cell(column=current_col_start + 1, row=i)  # lecturer cell
                    lec_cell.alignment = styles['wrap_text']
                    lec_cell.font = styles['default_font']

                    type_cell = ws.cell(column=current_col_start + 2, row=i)  # type cell
                    type_cell.alignment = styles['wrap_text']
                    type_cell.font = styles['default_font']

                    room_cell = ws.cell(column=current_col_start + 3, row=i)  # room cell
                    room_cell.alignment = styles['wrap_text']
                    room_cell.font = styles['default_font']

                for schedule in schedules:
                    k = (schedule.week_day - 1) * 6 + 6  # week day start row
                    j = k + schedule.pair_num - 1  # current row

                    ws.cell(column=current_col_start,
                            row=j,
                            value=fill_cell(ws, current_col_start, j, schedule.lesson.subject) + star[schedule.repeat_option] + \
                                  (f"(1/{schedule.lesson.group.subgroups})" if schedule.subgroup else ""))

                    if schedule.lesson.lecturer:
                        n = schedule.lesson.lecturer.name
                        short_lecturer_name = u'{} {}. {}.'.format(
                            n[:n.find(' ')],
                            n[n.find(' ') + 1],
                            n[n.rfind(' ') + 1]
                        )
                    else:
                        short_lecturer_name = '-'

                    ws.cell(column=current_col_start + 1,
                            row=j,
                            value=fill_cell(ws, current_col_start + 1, j, short_lecturer_name))

                    ws.cell(column=current_col_start + 2,
                            row=j,
                            value=fill_cell(ws, current_col_start + 2, j, schedule_types[schedule.type]))

                    ws.cell(column=current_col_start + 3,
                            row=j,
                            value=fill_cell(ws, current_col_start + 3, j, schedule.room.num))

    workbook.remove_sheet(workbook.get_sheet_by_name('Лист1'))
    workbook.save("imi_schedule.xlsx")
