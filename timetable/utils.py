import datetime
import os

from django.db.models import Q

from timetable import models
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, numbers, Side
from openpyxl.utils import get_column_letter


def generate_events(schedule):
    lesson_id = schedule.lesson.id

    semester_begin = models.Lesson.objects.get(pk=lesson_id).semester.study_start
    semester_end = models.Lesson.objects.get(pk=lesson_id).semester.study_end

    """Заполнение event"""
    # startpoint_date = datetime.date.today()
    startpoint_date = semester_begin - datetime.timedelta(days=1)
    if startpoint_date < semester_begin:
        startpoint_date = semester_begin
    if startpoint_date.isoweekday() > schedule.week_day:
        event_date = startpoint_date + datetime.timedelta(days=(7 - startpoint_date.isoweekday())) + \
                     datetime.timedelta(days=schedule.week_day)
    else:
        event_date = startpoint_date + datetime.timedelta(
            days=(schedule.week_day - startpoint_date.isoweekday()))
    weeks_count = (semester_end - event_date).days // 7 + 1
    pairtime_begin = {
        1: datetime.time(8, 0),
        2: datetime.time(9, 50),
        3: datetime.time(11, 40),
        4: datetime.time(14, 00),
        5: datetime.time(15, 50),
        6: datetime.time(17, 40)
    }
    pairtime_end = {
        1: datetime.time(9, 35),
        2: datetime.time(11, 25),
        3: datetime.time(13, 15),
        4: datetime.time(15, 35),
        5: datetime.time(17, 25),
        6: datetime.time(19, 15)
    }
    step = 2
    if schedule.repeat_option == 1:
        if event_date.isocalendar().week % 2 == 0:
            event_date += datetime.timedelta(days=7)
    elif schedule.repeat_option == 2:
        if event_date.isocalendar().week % 2 == 1:
            event_date += datetime.timedelta(days=7)
    else:
        step = 1

    # related_events = models.Event.objects.filter(schedule=schedule.id).filter(begin__gt=datetime.datetime.now())
    related_events = models.Event.objects.filter(schedule=schedule.id).filter(begin__gt=startpoint_date)
    if related_events.exists():
        related_events.delete()

    if schedule.subgroup and schedule.lesson.group.subgroups == None:
        raise Exception("У группы не заданы количество подгрупп")

    for i in range(0, weeks_count, step):
        event = models.Event()
        event.begin = datetime.datetime.combine(event_date + datetime.timedelta(days=7 * i),
                                                pairtime_begin[schedule.pair_num])
        event.end = datetime.datetime.combine(event_date + datetime.timedelta(days=7 * i),
                                                           pairtime_end[schedule.pair_num])
        event.lesson = schedule.lesson
        event.room = schedule.room
        event.type = schedule.type
        event.schedule = schedule
        event.save()


def get_week_events(viewset, request):
    start_date = request.GET['start_date']
    end_date = request.GET['end_date']

    if 'get_by' not in request.GET:
        return Response([])

    if request.GET['get_by'] == 'group':
        group_id = request.GET['param_id']
        events = models.Event.objects.filter(
            lesson__group__id=group_id,
            begin__range=(start_date, end_date)
        )
        if not events.exists():
            return Response({'error': 'Нет занятий либо такой группы не существует'}, status.HTTP_404_NOT_FOUND)
    elif request.GET['get_by'] == 'lecturer':
        lecturer_id = request.GET['param_id']
        events = models.Event.objects.filter(
            lesson__lecturer__id=lecturer_id,
            begin__range=(start_date, end_date)
        )
        if not events.exists():
            return Response({'error': 'Нет занятий либо такого преподавателя не существует'}, status.HTTP_404_NOT_FOUND)
    elif request.GET['get_by'] == 'room':
        room_id = request.GET['param_id']
        events = models.Event.objects.filter(
            room__id__iexact=room_id,
            begin__range=(start_date, end_date)
        )
        if not events.exists():
            return Response({'error': 'Нет занятий либо такого преподавателя не существует'}, status.HTTP_404_NOT_FOUND)
    else:
        return Response({'error': 'Неверный параметр {\'get_by\'} запроса'})
    serializer = viewset.get_serializer(events, many=True)
    return Response(serializer.data)


def get_event_date_and_weeks_count(group_id, week_day):
    """
        На выход дает:
                        - event_date - стартовую дату event-а
                        - weeks_count - количество недель в семестре
    """
    result = dict()

    semester_begin = models.Semester.objects.get(group_id=group_id).study_start
    semester_end = models.Semester.objects.get(group_id=group_id).study_end

    startpoint_date = semester_begin - datetime.timedelta(days=1)

    if startpoint_date < semester_begin:
        startpoint_date = semester_begin

    if startpoint_date.isoweekday() > week_day:
        event_date = startpoint_date + datetime.timedelta(days=(7 - startpoint_date.isoweekday())) + \
                     datetime.timedelta(days=week_day)
    else:
        event_date = startpoint_date + datetime.timedelta(
            days=(week_day - startpoint_date.isoweekday()))

    weeks_count = (semester_end - event_date).days // 7 + 1

    result['start_event_date'] = event_date
    result['weeks_count'] = weeks_count

    return result


def get_pairs_begin_end(pair_num):
    result = dict()
    pairtime_begin = {
        1: datetime.time(8, 0),
        2: datetime.time(9, 50),
        3: datetime.time(11, 40),
        4: datetime.time(14, 00),
        5: datetime.time(15, 50),
        6: datetime.time(17, 40)
    }
    pairtime_end = {
        1: datetime.time(9, 35),
        2: datetime.time(11, 25),
        3: datetime.time(13, 15),
        4: datetime.time(15, 35),
        5: datetime.time(17, 25),
        6: datetime.time(19, 15)
    }

    result['begin'] = pairtime_begin[pair_num]
    result['end'] = pairtime_end[pair_num]

    return result


def get_available_rooms(group_id, week_day, pair_num, repeat_option):
    rooms = models.Room.objects.all()

    event_date = get_event_date_and_weeks_count(group_id, week_day)['start_event_date']
    weeks_count = get_event_date_and_weeks_count(group_id, week_day)['weeks_count']

    pairtime_begin, pairtime_end = get_pairs_begin_end(pair_num)['begin'], get_pairs_begin_end(pair_num)[
        'end']

    if repeat_option == 0:
        step = 1
    else:
        step = 2

    if (repeat_option == 1 and event_date.isocalendar().week % 2 == 0) or \
            (repeat_option == 2 and event_date.isocalendar().week % 2 == 1):
        event_date += datetime.timedelta(days=7)

    for i in range(0, weeks_count, step):
        begin = datetime.datetime.combine(event_date + datetime.timedelta(days=7 * i), pairtime_begin)
        end = datetime.datetime.combine(event_date + datetime.timedelta(days=7 * i), pairtime_end)

        events = models.Event.objects.filter(Q(begin__range=(begin, end))
                                             |
                                             Q(end__range=(begin, end)))
        rooms = rooms.exclude(event__in=events)

    return rooms


def schedule_to_excel():

    app_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(app_dir, 'schedule.xlsx')
    workbook = load_workbook(template)
    ws = workbook.active

    groups = models.Group.objects.filter(name__contains="-18")
    # group = models.Group.objects.get(pk=group_id)

    star = {
        0: "",
        1: "*",
        2: "**",
    }

    schedule_types = {
        "LEC": "Лек",
        "PRA": "Пр",
        "LAB": "Лаб",
    }

    """ Styles """
    wrap_text = Alignment(horizontal="center",
                          vertical="center",
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=False,
                          indent=0)

    default_font = Font(name="Times New Roman",
                      size=14)

    header_font = Font(name="Times New Roman",
                       size=14,
                       bold="b")


    c = 0
    for group in groups:
        c += 1
        semester = models.Semester.objects.get(group__id=group.id)

        current_col_start = int(3 + (c - 1) * 4)

        top_range = ws[get_column_letter(current_col_start) + "3" : get_column_letter(current_col_start + 3) + "3"]
        bottom_range = ws[get_column_letter(current_col_start) + "41" : get_column_letter(current_col_start + 3) + "41"]
        left_range = ws[get_column_letter(current_col_start) + "3" : get_column_letter(current_col_start) + "41"]
        right_range = ws[get_column_letter(current_col_start + 3) + "3" : get_column_letter(current_col_start + 3) + "41"]


        rrange = ws[get_column_letter(current_col_start) + "3" : get_column_letter(current_col_start + 3) + "41"]

        black_medium_side = Side(border_style="medium", color="000000")

        for row in top_range:
            for cell in row:
                cell.border = Border(**(cell.border.__dict__ | {'top': black_medium_side}))

        for row in bottom_range:
            for cell in row:
                cell.border = Border(**(cell.border.__dict__ | {'bottom': black_medium_side}))

        for row in left_range:
            for cell in row:
                cell.border = Border(
                    **(cell.border.__dict__ | {'left': black_medium_side}))

        for row in right_range:
            for cell in row:
                cell.border = Border(**(cell.border.__dict__ | {'right': black_medium_side}))


        ws.column_dimensions[get_column_letter(current_col_start)].width = 35
        ws.column_dimensions[get_column_letter(current_col_start + 1)].width = 19
        ws.column_dimensions[get_column_letter(current_col_start + 2)].width = 14
        ws.column_dimensions[get_column_letter(current_col_start + 3)].width = 14

        gdescell = ws.cell(column=current_col_start, row=3) # group description cell
        gdescell.value = "Наименование группы"
        gdescell.font = default_font

        scdescell = ws.cell(column=current_col_start + 1, row=3) # students count description cell
        scdescell.value = "Общее число обучающихся, чел."
        scdescell.font = default_font

        gcell = ws.cell(column=current_col_start, row=4) # group name cell
        gcell.value = group.name
        gcell.font = header_font
        gcell.alignment = wrap_text

        sbcell = ws.cell(column=current_col_start + 2, row=4) # semester begin cell
        sbcell.value = semester.study_start
        sbcell.font = header_font
        sbcell.alignment = wrap_text
        sbcell.number_format = numbers.BUILTIN_FORMATS[16]

        secell = ws.cell(column=current_col_start + 3, row=4)
        secell.value = semester.study_end
        secell.font = header_font
        secell.alignment = wrap_text
        secell.number_format = numbers.BUILTIN_FORMATS[16]

        disdescell = ws.cell(column=current_col_start, row=5) # discipline desc cell
        disdescell.value = "Дисциплина"
        disdescell.alignment = wrap_text
        disdescell.font = default_font

        lecdescell = ws.cell(column=current_col_start + 1, row=5) # lecturer desc cell
        lecdescell.value = "ФИО преподавателя"
        lecdescell.alignment = wrap_text
        lecdescell.font = default_font

        typedescell = ws.cell(column=current_col_start + 2, row=5)  # type desc cell
        typedescell.value = "Вид учебного занятия"
        typedescell.alignment = wrap_text
        typedescell.font = default_font

        roomdescell = ws.cell(column=current_col_start + 3, row=5)  # room desc cell
        roomdescell.value = "ауд."
        roomdescell.alignment = wrap_text
        roomdescell.font = default_font

        schedules = models.Schedule.objects.filter(lesson__group__id=group.id).order_by("week_day", "pair_num")


        def fill_cell(column, j, content):
            return  ((str(ws.cell(column=column, row=j).value) + "\n")
                     if ws.cell(column=column, row=j).value != None else "") + content

        for i in range(5, 42):
            discell = ws.cell(column=current_col_start, row=i) # discipline cell
            discell.alignment = wrap_text
            discell.font = default_font

            lecell = ws.cell(column=current_col_start + 1, row=i) # lecturer cell
            lecell.alignment = wrap_text
            lecell.font = default_font

            typecell = ws.cell(column=current_col_start + 2, row=i) # type cell
            typecell.alignment = wrap_text
            typecell.font = default_font

            roomcell = ws.cell(column=current_col_start + 3, row=i) # room cell
            roomcell.alignment = wrap_text
            roomcell.font = default_font



        for schedule in schedules:
            k = (schedule.week_day - 1) * 6 + 6 # week day start row
            j = k + schedule.pair_num - 1 # current row

            ws.cell(column=current_col_start,
                    row=j,
                    value=fill_cell(current_col_start, j, schedule.lesson.subject) + star[schedule.repeat_option] + \
                          (f"(1/{schedule.lesson.group.subgroups})" if schedule.subgroup else ""))

            n = schedule.lesson.lecturer.name
            short_lecturer_name = u'{} {}. {}.'.format(
                n[:n.find(' ')],
                n[n.find(' ') + 1],
                n[n.rfind(' ') + 1]
            )

            ws.cell(column=current_col_start + 1,
                    row=j,
                    value=fill_cell(current_col_start + 1, j, short_lecturer_name))

            ws.cell(column=current_col_start + 2,
                    row=j,
                    value=fill_cell(current_col_start + 2, j, schedule_types[schedule.type]))

            ws.cell(column=current_col_start + 3,
                    row=j,
                    value=fill_cell(current_col_start + 3, j, schedule.room.num))

    workbook.save("imi_schedule.xlsx")
